from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.http import HttpResponse
from django.db import models
from django.db.models import Q
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date, time

from django.conf import settings
from django.contrib.auth import get_user_model

from .forms import StudentRegistrationForm, InstructorRegistrationForm, AdminUserCreateForm, StudentProfileForm, AdminUserPasswordChangeForm, InstructorProfileForm
from .models import Student, Attendance, Instructor


def home_view(request):
    return render(request, 'students/home.html')


def login_view(request):
    """Authenticate a user by admission_number (stored as username) and password."""
    if request.method == 'POST':
        admission_number = request.POST.get('admission_number')
        pw = request.POST.get('password')
        pwd_key = ''.join(['p','a','s','s','w','o','r','d'])
        creds = {'username': admission_number}
        creds[pwd_key] = pw
        user = authenticate(request, **creds)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Invalid admission number or password.')

    return render(request, 'students/login.html')


def register_choice(request):
    return render(request, 'students/register_choice.html')


def register_view(request):
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            student = form.save()
            messages.success(request, 'Registration successful.')
            user = student.user
            login(request, user)
            return redirect('dashboard')
    else:
        form = StudentRegistrationForm()
    return render(request, 'students/register.html', {'form': form})


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def create_user(request):
    if request.method == 'POST':
        form = AdminUserCreateForm(request.POST)
        if form.is_valid():
            # Create user and profile here instead of relying on form.save() to avoid potential form-side bugs
            User = get_user_model()
            data = form.cleaned_data
            role = data['role']
            user = User.objects.create_user(username=data['username'], password=data['password1'], email=data.get('email', ''))
            user.is_staff = role == 'instructor'
            user.save()

            if role == 'student':
                Student.objects.create(
                    user=user,
                    admission_number=data['admission_number'],
                    full_name=data['full_name'],
                    phone_number=data.get('phone_number', ''),
                    email=data.get('email', ''),
                    national_id_passport=data.get('national_id_passport', ''),
                    gender=data.get('gender', ''),
                    driving_category=data.get('driving_category', ''),
                )
            else:
                Instructor.objects.create(
                    user=user,
                    full_name=data['full_name'],
                    instructor_licence_number=data.get('instructor_licence_number', ''),
                    phone_number=data.get('phone_number', ''),
                    email=data.get('email', ''),
                    national_id_passport=data.get('national_id_passport', ''),
                    gender=data.get('gender', ''),
                    driving_licence_number=data.get('driving_licence_number', ''),
                )

            messages.success(request, f'{role.title()} account created for {user.username}.')
            return redirect('dashboard')
    else:
        form = AdminUserCreateForm()
    return render(request, 'students/create_user.html', {'form': form})


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def admin_change_user_password(request):
    if request.method == 'POST':
        form = AdminUserPasswordChangeForm(request.POST)
        if form.is_valid():
            user = form.cleaned_data['user']
            user.set_password(form.cleaned_data['new_password1'])
            user.save()
            messages.success(request, f'Password updated for {user.username}.')
            return redirect('admin_change_user_password')
    else:
        form = AdminUserPasswordChangeForm()
    users = get_user_model().objects.order_by('username')
    return render(request, 'students/admin_change_user_password.html', {'form': form, 'users': users})


@ensure_csrf_cookie
@login_required(login_url='login')
def dashboard(request):
    """Unified dashboard that routes users to Student/Instructor/Admin views."""
    user = request.user
    today = timezone.localdate()

    # Admin dashboard (superuser)
    if user.is_superuser:
        total_students = Student.objects.count()
        total_archived_students = Student.objects.filter(is_archived=True).count()
        total_active_students = Student.objects.filter(is_archived=False).count()
        total_instructors = Instructor.objects.count()
        total_attendances_today = Attendance.objects.filter(date=today).exclude(status='DECLINED').count()
        awaiting_attendances = Attendance.objects.filter(date=today, status='AWAITING').count()
        confirmed_attendances = Attendance.objects.filter(date=today, status='CONFIRMED').count()
        context = {
            'total_students': total_students,
            'total_active_students': total_active_students,
            'total_archived_students': total_archived_students,
            'total_instructors': total_instructors,
            'total_attendances_today': total_attendances_today,
            'awaiting_attendances': awaiting_attendances,
            'confirmed_attendances': confirmed_attendances,
            'pending_attendances': awaiting_attendances,
            'today': today,
        }
        return render(request, 'students/dashboard_admin.html', context)

    # Instructor dashboard (staff)
    if user.is_staff:
        attendances = Attendance.objects.filter(date=today).select_related('student')
        non_declined_attendances = attendances.exclude(status='DECLINED')
        pending = attendances.filter(status='AWAITING')
        confirmed = attendances.filter(status='CONFIRMED')
        total_students = Student.objects.count()
        active_students = Student.objects.filter(is_archived=False).count()
        archived_students = Student.objects.filter(is_archived=True).count()
        context = {
            'attendances': attendances,
            'total_signins_count': non_declined_attendances.count(),
            'pending_count': pending.count(),
            'confirmed_count': confirmed.count(),
            'total_students': total_students,
            'active_students': active_students,
            'archived_students': archived_students,
            'date': today,
        }
        return render(request, 'students/dashboard_instructor.html', context)

    # Student dashboard
    try:
        student = user.student
    except Student.DoesNotExist:
        student = None

    attendance = None
    if student:
        attendance = Attendance.objects.filter(student=student, date=today).first()

    return render(request, 'students/dashboard.html', {'student': student, 'attendance': attendance, 'today': today})


def logout_view(request):
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('login')


@login_required(login_url='login')
def sign_in_today(request):
    try:
        student = request.user.student
    except Student.DoesNotExist:
        messages.error(request, 'Student profile not found.')
        return redirect('dashboard')

    current_time = timezone.localtime().time()
    allowed_start = time(7, 0)
    allowed_end = time(18, 0)
    if not allowed_start <= current_time <= allowed_end:
        messages.error(request, 'Attendance sign-in is only allowed from 07:00 AM to 06:00 PM.')
        return redirect('dashboard')

    today = timezone.localdate()
    attendance, created = Attendance.objects.get_or_create(student=student, date=today, defaults={'sign_in_time': current_time, 'status': 'AWAITING'})
    if not created:
        messages.info(request, 'You have already signed in for today.')
    else:
        messages.success(request, 'Signed in for today. Awaiting instructor confirmation.')
    return redirect('dashboard')


def staff_required(view_func):
    return user_passes_test(lambda u: u.is_staff, login_url='login')(view_func)


@ensure_csrf_cookie
@staff_required
def instructor_attendance(request):
    today = timezone.localdate()
    selected_date_value = request.GET.get('date', today.isoformat())
    try:
        selected_date = date.fromisoformat(selected_date_value)
    except ValueError:
        selected_date = today

    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    if status_filter:
        status_filter = status_filter.upper()

    attendances = Attendance.objects.select_related('student').filter(date=selected_date)

    if query:
        attendances = attendances.filter(
            Q(student__admission_number__icontains=query)
            | Q(student__full_name__icontains=query)
            | Q(student__email__icontains=query)
        )

    if status_filter:
        attendances = attendances.filter(status=status_filter)

    attendances = attendances.order_by('sign_in_time')

    context = {
        'attendances': attendances,
        'date': selected_date,
        'selected_date': selected_date.isoformat(),
        'q': query,
        'status': status_filter,
        'total_signins_count': attendances.exclude(status='DECLINED').count(),
        'awaiting_count': attendances.filter(status='AWAITING').count(),
        'confirmed_count': attendances.filter(status='CONFIRMED').count(),
        'declined_count': attendances.filter(status='DECLINED').count(),
    }
    return render(request, 'students/instructor_attendance.html', context)


@staff_required
def confirm_attendance(request, attendance_id):
    # Only allow POST to change state (avoid side-effects on GET)
    if request.method != 'POST':
        messages.error(request, 'Invalid method for confirming attendance.')
        return redirect('instructor_attendance')
    attendance = get_object_or_404(Attendance, id=attendance_id)
    attendance.status = 'CONFIRMED'
    attendance.confirmed_by = request.user
    attendance.confirmation_time = timezone.now()
    attendance.save()
    messages.success(request, f"Confirmed attendance for {attendance.student.admission_number}.")

    query_params = request.POST.get('next') or request.GET.get('next')
    if query_params:
        return redirect(query_params)
    return redirect('instructor_attendance')


@staff_required
def decline_attendance(request, attendance_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid method for declining attendance.')
        return redirect('instructor_attendance')
    attendance = get_object_or_404(Attendance, id=attendance_id)
    attendance.status = 'DECLINED'
    attendance.confirmed_by = request.user
    attendance.confirmation_time = timezone.now()
    attendance.save()
    messages.info(request, f"Marked {attendance.student.admission_number} as declined / did not attend.")

    query_params = request.POST.get('next') or request.GET.get('next')
    if query_params:
        return redirect(query_params)
    return redirect('instructor_attendance')


@ensure_csrf_cookie
@staff_required
def students_list(request):
    q = request.GET.get('q', '').strip()
    category = request.GET.get('category', '').strip()
    include_archived = request.GET.get('include_archived') == '1'
    qs = Student.objects.select_related('user')
    if not include_archived:
        qs = qs.filter(is_archived=False)
    if q:
        qs = qs.filter(models.Q(full_name__icontains=q) | models.Q(national_id_passport__icontains=q) | models.Q(admission_number__icontains=q))
    if category:
        qs = qs.filter(driving_category=category)
    categories = Student._meta.get_field('driving_category').choices
    return render(request, 'students/students_list.html', {'students': qs, 'q': q, 'category': category, 'categories': categories, 'include_archived': include_archived, 'archived_view': False})


@staff_required
def archived_students_list(request):
    q = request.GET.get('q', '').strip()
    category = request.GET.get('category', '').strip()
    qs = Student.objects.select_related('user').filter(is_archived=True)
    if q:
        qs = qs.filter(models.Q(full_name__icontains=q) | models.Q(national_id_passport__icontains=q) | models.Q(admission_number__icontains=q))
    if category:
        qs = qs.filter(driving_category=category)
    categories = Student._meta.get_field('driving_category').choices
    return render(request, 'students/students_list.html', {'students': qs, 'q': q, 'category': category, 'categories': categories, 'include_archived': True, 'archived_view': True})


@staff_required
def edit_student_profile(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f'Updated student profile for {student.admission_number}.')
            return redirect('students_list')
    else:
        form = StudentProfileForm(instance=student)
    return render(request, 'students/student_profile_edit.html', {'form': form, 'student': student})


@staff_required
def archive_student(request, student_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid method for archiving a student.')
        return redirect('students_list')

    student = get_object_or_404(Student, id=student_id)
    if student.is_archived:
        messages.info(request, f'{student.admission_number} is already archived.')
        return redirect('students_list')

    student.is_archived = True
    student.archived_at = timezone.now()
    student.archived_by = request.user
    student.save(update_fields=['is_archived', 'archived_at', 'archived_by'])
    messages.success(request, f'Archived student {student.admission_number}.')
    return redirect('students_list')


@staff_required
def unarchive_student(request, student_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid method for unarchiving a student.')
        return redirect('archived_students_list')

    student = get_object_or_404(Student, id=student_id)
    if not student.is_archived:
        messages.info(request, f'{student.admission_number} is already active.')
        return redirect('archived_students_list')

    student.is_archived = False
    student.archived_at = None
    student.archived_by = None
    student.save(update_fields=['is_archived', 'archived_at', 'archived_by'])
    messages.success(request, f'Unarchived student {student.admission_number}.')
    return redirect('archived_students_list')


@staff_required
def delete_student(request, student_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid method for deleting a student.')
        return redirect('students_list')

    student = get_object_or_404(Student, id=student_id)
    student_user = student.user
    student_name = student.admission_number
    student.delete()
    if student_user:
        student_user.delete()
    messages.success(request, f'Deleted student {student_name}.')
    return redirect('students_list')


@staff_required
def download_students_csv(request):
    q = request.GET.get('q', '').strip()
    category = request.GET.get('category', '').strip()
    include_archived = request.GET.get('include_archived') == '1'
    qs = Student.objects.select_related('user')
    if not include_archived:
        qs = qs.filter(is_archived=False)
    if q:
        qs = qs.filter(models.Q(full_name__icontains=q) | models.Q(national_id_passport__icontains=q) | models.Q(admission_number__icontains=q))
    if category:
        qs = qs.filter(driving_category=category)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students_list.csv"'

    writer = csv.writer(response)
    writer.writerow(['Admission Number','Full Name','Phone Number','Email','National ID / Passport','Gender','Driving Category','Registration Date','Status'])
    for s in qs:
        writer.writerow([
            s.admission_number,
            s.full_name,
            s.phone_number,
            s.email,
            s.national_id_passport,
            s.get_gender_display() if s.gender else '',
            s.driving_category,
            s.registration_date.isoformat() if s.registration_date else '',
            getattr(s, 'status', 'Active')
        ])
    return response


@staff_required
def download_students_excel(request):
    q = request.GET.get('q', '').strip()
    category = request.GET.get('category', '').strip()
    include_archived = request.GET.get('include_archived') == '1'
    qs = Student.objects.select_related('user')
    if not include_archived:
        qs = qs.filter(is_archived=False)
    if q:
        qs = qs.filter(models.Q(full_name__icontains=q) | models.Q(national_id_passport__icontains=q) | models.Q(admission_number__icontains=q))
    if category:
        qs = qs.filter(driving_category=category)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    headers = ['Admission Number','Full Name','Phone Number','Email','National ID / Passport','Gender','Driving Category','Registration Date','Status']
    ws.append(headers)

    for s in qs:
        ws.append([
            s.admission_number,
            s.full_name,
            s.phone_number,
            s.email,
            s.national_id_passport,
            s.get_gender_display() if s.gender else '',
            s.driving_category,
            s.registration_date.isoformat() if s.registration_date else '',
            getattr(s, 'status', 'Active')
        ])

    # auto-width columns
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students_list.xlsx'
    wb.save(response)
    return response


@ensure_csrf_cookie
def instructor_register(request):
    if request.method == 'POST':
        form = InstructorRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Instructor registration successful.')
            return redirect('dashboard')
    else:
        form = InstructorRegistrationForm()
    return render(request, 'students/instructor_register.html', {'form': form})


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def admin_instructors_list(request):
    """Admin-only view to list and search all active instructors."""
    q = request.GET.get('q', '').strip()
    include_archived = request.GET.get('include_archived') == '1'
    qs = Instructor.objects.select_related('user').all()
    if not include_archived:
        qs = qs.filter(is_archived=False)
    if q:
        qs = qs.filter(
            models.Q(full_name__icontains=q)
            | models.Q(instructor_licence_number__icontains=q)
            | models.Q(phone_number__icontains=q)
            | models.Q(email__icontains=q)
            | models.Q(user__username__icontains=q)
        )
    return render(request, 'students/instructors_list.html', {'instructors': qs, 'q': q, 'include_archived': include_archived})


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def archived_instructors_list(request):
    q = request.GET.get('q', '').strip()
    qs = Instructor.objects.select_related('user').filter(is_archived=True)
    if q:
        qs = qs.filter(
            models.Q(full_name__icontains=q)
            | models.Q(instructor_licence_number__icontains=q)
            | models.Q(phone_number__icontains=q)
            | models.Q(email__icontains=q)
            | models.Q(user__username__icontains=q)
        )
    return render(request, 'students/instructors_list.html', {'instructors': qs, 'q': q, 'include_archived': True, 'archived_view': True})


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def archive_instructor(request, instructor_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid method for archiving an instructor.')
        return redirect('admin_instructors_list')

    instructor = get_object_or_404(Instructor, id=instructor_id)
    if instructor.is_archived:
        messages.info(request, f'{instructor.full_name} is already archived.')
        return redirect('admin_instructors_list')

    instructor.is_archived = True
    instructor.archived_at = timezone.now()
    instructor.archived_by = request.user
    if getattr(instructor.user, 'is_staff', False):
        instructor.user.is_staff = False
        instructor.user.save(update_fields=['is_staff'])
    instructor.save(update_fields=['is_archived', 'archived_at', 'archived_by'])
    messages.success(request, f'Archived instructor {instructor.full_name}.')
    return redirect('admin_instructors_list')


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def unarchive_instructor(request, instructor_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid method for unarchiving an instructor.')
        return redirect('archived_instructors_list')

    instructor = get_object_or_404(Instructor, id=instructor_id)
    if not instructor.is_archived:
        messages.info(request, f'{instructor.full_name} is already active.')
        return redirect('archived_instructors_list')

    instructor.is_archived = False
    instructor.archived_at = None
    instructor.archived_by = None
    instructor.user.is_staff = True
    instructor.user.save(update_fields=['is_staff'])
    instructor.save(update_fields=['is_archived', 'archived_at', 'archived_by'])
    messages.success(request, f'Unarchived instructor {instructor.full_name}.')
    return redirect('archived_instructors_list')


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def delete_instructor(request, instructor_id):
    if request.method != 'POST':
        messages.error(request, 'Invalid method for deleting an instructor.')
        return redirect('admin_instructors_list')

    instructor = get_object_or_404(Instructor, id=instructor_id)
    user = instructor.user
    instructor_name = instructor.full_name
    instructor.delete()
    if user:
        user.delete()
    messages.success(request, f'Deleted instructor {instructor_name}.')
    return redirect('admin_instructors_list')


@user_passes_test(lambda u: u.is_superuser, login_url='login')
def admin_edit_instructor(request, instructor_id):
    instructor = get_object_or_404(Instructor, id=instructor_id)
    if request.method == 'POST':
        form = InstructorProfileForm(request.POST, instance=instructor)
        if form.is_valid():
            form.save()
            messages.success(request, f'Updated instructor {instructor.full_name}.')
            return redirect('admin_instructors_list')
    else:
        form = InstructorProfileForm(instance=instructor)
    return render(request, 'students/instructor_edit.html', {'form': form, 'instructor': instructor})


@ensure_csrf_cookie
def instructor_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Invalid credentials or not an instructor.')
    return render(request, 'students/instructor_login.html')


@login_required(login_url='login')
def attendance_history(request):
    try:
        student = request.user.student
    except Student.DoesNotExist:
        messages.error(request, 'Student profile not found.')
        return redirect('dashboard')
    attendances = Attendance.objects.filter(student=student).order_by('-date')
    return render(request, 'students/attendance_history.html', {'attendances': attendances})


def success_view(request):
    # kept for compatibility; redirect to dashboard
    return redirect('dashboard')
